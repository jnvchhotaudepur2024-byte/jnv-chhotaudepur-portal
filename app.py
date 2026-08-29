import io, os, re, json, sqlite3, hashlib, zipfile, shutil
from datetime import datetime
import pandas as pd
import numpy as np
import openpyxl
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

APP_TITLE = 'PM SHRI JNV CHHOTAUDEPUR - RESULT PORTAL'
DB_FILE = 'school_database.db'
BACKUP_DIR = 'backups'
HISTORY_FILE = 'update_history.csv'
PHOTO_DIR = 'photos/students'
os.makedirs(BACKUP_DIR, exist_ok=True); os.makedirs(PHOTO_DIR, exist_ok=True)

# Subjects exactly aligned with the supplied Excel templates.
LOWER_SUBJECTS = ['English','Hindi','Gujarati','Mathematics','Science','Social Science']
IX_SUBJECTS = LOWER_SUBJECTS + ['3rd Language','Skill Subject']
XI_SUBJECTS = ['Subject-1','Subject-2','Subject-3','Subject-4','Subject-5','Subject-6','Additional Subject']

META = ['Class','Roll_No','GR_No','Student_Name','DOB','Gender','Mother_Name','Father_Name','Area','House','Mobile_No',
        'Class_Teacher','Attendance','Working_Days','Present_Days','Discipline','Skill_Course','Co_Scholastic',
        'Bagless_Days','Outstanding_Achievement','Remarks']

def norm(v):
    if v is None or (isinstance(v,float) and np.isnan(v)): return ''
    if isinstance(v, (datetime,)): return v.strftime('%d-%m-%Y')
    if hasattr(v,'strftime') and not isinstance(v,str):
        try: return v.strftime('%d-%m-%Y')
        except: pass
    s=str(v).strip()
    if re.fullmatch(r'-?\d+\.0+',s): s=s.split('.')[0]
    return s

def num(v):
    try:
        if v is None or str(v).strip()=='': return np.nan
        return float(v)
    except: return np.nan

def grade(p):
    if pd.isna(p): return ''
    p=float(p)
    return 'A1' if p>=91 else 'A2' if p>=81 else 'B1' if p>=71 else 'B2' if p>=61 else 'C1' if p>=51 else 'C2' if p>=41 else 'D' if p>=33 else 'E'

def read_values(path, sheet):
    return pd.read_excel(path, sheet_name=sheet, header=None, engine='openpyxl')

def find_class(wb, fallback=''):
    for s in ['School_Data','Sheet1 (3)']:
        if s in wb.sheetnames:
            ws=wb[s]
            for row in ws.iter_rows(min_row=1,max_row=min(ws.max_row,25),values_only=True):
                text=' '.join(norm(x) for x in row if x is not None)
                m=re.search(r'CLASS\s*[:\-]\s*(VI{0,2}|IX|XI|[6-9]|11)',text,re.I)
                if m:
                    x=m.group(1).upper(); return {'6':'VI','7':'VII','8':'VIII','9':'IX','11':'XI'}.get(x,x)
    if 'CLASS XI' in os.path.basename(fallback).upper(): return 'XI'
    if 'CLASS IX' in os.path.basename(fallback).upper(): return 'IX'
    return 'VI-VIII'

def import_template(path):
    """Read the actual Data/DataSheet sheet of the supplied JNV templates.
    Formula cells are not trusted: the Python portal recalculates totals/ranks.
    """
    wb=openpyxl.load_workbook(path,data_only=True,read_only=True)
    cls=find_class(wb,path)
    sheet = 'DataSheet' if 'DataSheet' in wb.sheetnames else 'DATA'
    ws=wb[sheet]
    rows=list(ws.iter_rows(values_only=True))
    if len(rows)<6: return pd.DataFrame()
    # The templates use row 5 as numeric column labels and data starts at row 6.
    records=[]
    if cls=='XI':
        # metadata cols 1-10, 7 subject blocks, each block 21 cols.
        starts=[11,32,53,74,95,116,137]
        for r in range(5,len(rows)):
            vals=rows[r]
            roll=norm(vals[0] if len(vals)>0 else '')
            if not roll or roll.lower()=='none': continue
            rec={'Class':'XI','Roll_No':roll,'GR_No':norm(vals[1]),'Student_Name':norm(vals[2]),'DOB':norm(vals[3]),
                 'House':norm(vals[4]),'Gender':norm(vals[5]),'Mother_Name':norm(vals[6]),'Father_Name':norm(vals[7]),
                 'Area':norm(vals[8]),'Mobile_No':norm(vals[9])}
            total=0.0; max_total=0.0
            breakdown={}
            for i,start in enumerate(starts):
                # block: T1 at start..start+7, T2 at start+8..start+15, grand at +16,
                # theory/practical at +17/+18, combined 100 at +19, grade +20
                if i>=len(XI_SUBJECTS): continue
                name=XI_SUBJECTS[i]
                t1=num(vals[start-1+7] if len(vals)>start+6 else np.nan) # zero-based +7
                t2=num(vals[start-1+15] if len(vals)>start+14 else np.nan)
                combined=num(vals[start-1+19] if len(vals)>start+18 else np.nan)
                if pd.isna(combined):
                    combined = (0.4*t1+0.6*t2) if not pd.isna(t1) and not pd.isna(t2) else (t1 if not pd.isna(t1) else t2)
                rec[name]=combined
                total += 0 if pd.isna(combined) else combined; max_total += 100 if not pd.isna(combined) else 0
                breakdown[name]={'term1':None if pd.isna(t1) else t1,'term2':None if pd.isna(t2) else t2,
                                 'combined':None if pd.isna(combined) else combined,
                                 'theory':None if pd.isna(num(vals[start-1+17])) else num(vals[start-1+17]),
                                 'practical':None if pd.isna(num(vals[start-1+18])) else num(vals[start-1+18])}
            rec.update({'Attendance':norm(vals[170]) if len(vals)>170 else '', 'Working_Days':norm(vals[170]) if len(vals)>170 else '',
                        'Present_Days':norm(vals[171]) if len(vals)>171 else '', 'Outstanding_Achievement':norm(vals[174]) if len(vals)>174 else '',
                        'Remarks':norm(vals[175]) if len(vals)>175 else '', 'Skill_Course':'', 'Discipline':'', 'Co_Scholastic':'', 'Bagless_Days':''})
            rec['Total_Marks']=round(total,2); rec['Max_Marks']=max_total; rec['Percentage']=round(total/max_total*100,2) if max_total else 0
            rec['Subject_Data']=json.dumps({k:rec.get(k) for k in XI_SUBJECTS},ensure_ascii=False); rec['Breakdown']=json.dumps(breakdown,ensure_ascii=False)
            records.append(rec)
    else:
        starts=[10,33,56,79,102,125]
        subjects=IX_SUBJECTS if cls=='IX' else LOWER_SUBJECTS
        for r in range(5,len(rows)):
            vals=rows[r]; roll=norm(vals[0] if len(vals)>0 else '')
            if not roll: continue
            rec={'Class':cls,'Roll_No':roll,'GR_No':norm(vals[1]),'Student_Name':norm(vals[2]),'DOB':norm(vals[3]),
                 'Gender':norm(vals[4]),'Mother_Name':norm(vals[5]),'Father_Name':norm(vals[6]),'Area':norm(vals[7]),'House':norm(vals[8])}
            # If VIII/IX has extra skill block, read the first 6 academic subjects here.
            total=0; max_total=0; breakdown={}
            for i,start in enumerate(starts):
                if i>=len(subjects): break
                name=subjects[i]
                t1=num(vals[start-1+8] if len(vals)>start+7 else np.nan)
                t2=num(vals[start-1+18] if len(vals)>start+17 else np.nan)
                grand=num(vals[start-1+20] if len(vals)>start+19 else np.nan)
                if pd.isna(grand): grand=0.4*t1+0.6*t2 if not pd.isna(t1) and not pd.isna(t2) else (t1 if not pd.isna(t1) else t2)
                rec[name]=grand
                if not pd.isna(grand): total+=grand; max_total+=100
                breakdown[name]={'term1':None if pd.isna(t1) else t1,'term2':None if pd.isna(t2) else t2,'combined':None if pd.isna(grand) else grand}
            # IX skill subject lives in a separate block; attempt to read it from the named HY/Y sheets.
            if cls=='IX' and 'Skill Subject Consolidated' in wb.sheetnames:
                sk=list(wb['Skill Subject Consolidated'].iter_rows(values_only=True))
                if r < len(sk):
                    sv=sk[r]
                    if len(sv)>6:
                        rec['Skill Subject']=num(sv[6])
                        if not pd.isna(rec['Skill Subject']): total+=rec['Skill Subject']; max_total+=100
            rec.update({'Mobile_No':'','Class_Teacher':'','Attendance':norm(vals[165]) if len(vals)>165 else '',
                        'Working_Days':norm(vals[165]) if len(vals)>165 else '', 'Present_Days':norm(vals[166]) if len(vals)>166 else '',
                        'Discipline':norm(vals[160]) if len(vals)>160 else '', 'Skill_Course':'','Co_Scholastic':'','Bagless_Days':'',
                        'Outstanding_Achievement':norm(vals[167]) if len(vals)>167 else '', 'Remarks':norm(vals[168]) if len(vals)>168 else ''})
            rec['Total_Marks']=round(total,2); rec['Max_Marks']=max_total; rec['Percentage']=round(total/max_total*100,2) if max_total else 0
            rec['Subject_Data']=json.dumps({k:rec.get(k) for k in subjects},ensure_ascii=False); rec['Breakdown']=json.dumps(breakdown,ensure_ascii=False)
            records.append(rec)
    df=pd.DataFrame(records)
    if not df.empty:
        df['Class_Rank']=df.groupby('Class')['Total_Marks'].rank(method='min',ascending=False).astype(int)
    return df

def init_db():
    con=sqlite3.connect(DB_FILE)
    con.execute('''CREATE TABLE IF NOT EXISTS students (key TEXT PRIMARY KEY, data TEXT NOT NULL, updated_at TEXT NOT NULL)''')
    con.execute('''CREATE TABLE IF NOT EXISTS history (id INTEGER PRIMARY KEY AUTOINCREMENT, key TEXT, action TEXT, old_data TEXT, new_data TEXT, changed_at TEXT)''')
    con.commit(); con.close()
init_db()

def key_for(r): return f"{norm(r.get('Class'))}|{norm(r.get('Roll_No'))}"

def db_load():
    con=sqlite3.connect(DB_FILE); rows=con.execute('SELECT data FROM students').fetchall(); con.close()
    return pd.DataFrame([json.loads(x[0]) for x in rows]) if rows else pd.DataFrame()

def db_merge(df):
    if df.empty: return
    con=sqlite3.connect(DB_FILE)
    now=datetime.now().isoformat(timespec='seconds')
    for _,row in df.iterrows():
        new={k:(None if pd.isna(v) else v) for k,v in row.to_dict().items()}
        k=key_for(new); oldrow=con.execute('SELECT data FROM students WHERE key=?',(k,)).fetchone()
        if oldrow:
            old=json.loads(oldrow[0]); merged=old.copy()
            # New non-empty Excel values replace old values. Empty cells never erase an existing value.
            for a,v in new.items():
                if v is not None and str(v).strip()!='': merged[a]=v
            con.execute('INSERT INTO history(key,action,old_data,new_data,changed_at) VALUES(?,?,?,?,?)',(k,'UPDATE',json.dumps(old,default=str),json.dumps(merged,default=str),now))
        else:
            merged=new
            con.execute('INSERT INTO history(key,action,old_data,new_data,changed_at) VALUES(?,?,?,?,?)',(k,'INSERT','',json.dumps(merged,default=str),now))
        con.execute('INSERT OR REPLACE INTO students(key,data,updated_at) VALUES(?,?,?)',(k,json.dumps(merged,default=str,ensure_ascii=False),now))
    con.commit(); con.close()

def backup_db():
    if os.path.exists(DB_FILE): shutil.copy2(DB_FILE,os.path.join(BACKUP_DIR,'school_database_'+datetime.now().strftime('%Y%m%d_%H%M%S')+'.db'))

def make_report(student, history_df=None):
    buf=io.BytesIO(); doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=18,leftMargin=18,topMargin=16,bottomMargin=16)
    styles=getSampleStyleSheet(); title=ParagraphStyle('title',parent=styles['Title'],fontSize=13,leading=15,alignment=1,textColor=colors.HexColor('#003366')); small=ParagraphStyle('small',parent=styles['Normal'],fontSize=7,leading=8); center=ParagraphStyle('center',parent=small,alignment=1)
    story=[Paragraph('PM SHRI SCHOOL JAWAHAR NAVODAYA VIDYALAYA CHHOTAUDEPUR',title),Paragraph('STUDENT REPORT CARD',ParagraphStyle('sub',parent=title,fontSize=10,textColor=colors.HexColor('#B22222'))),Spacer(1,6)]
    info=[['Class',norm(student.get('Class')),'Roll No.',norm(student.get('Roll_No')),'G.R. No.',norm(student.get('GR_No'))],['Student Name',norm(student.get('Student_Name')),'DOB',norm(student.get('DOB')),'Gender',norm(student.get('Gender'))],['Father Name',norm(student.get('Father_Name')),'Mother Name',norm(student.get('Mother_Name')),'House',norm(student.get('House'))],['Area',norm(student.get('Area')),'Mobile',norm(student.get('Mobile_No')),'Class Teacher',norm(student.get('Class_Teacher'))]]
    t=Table(info,colWidths=[55,175,50,110,55,90]); t.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.4,colors.grey),('FONTNAME',(0,0),(-1,-1),'Helvetica'),('FONTSIZE',(0,0),(-1,-1),7),('BACKGROUND',(0,0),(0,-1),colors.HexColor('#E8F1F8')),('BACKGROUND',(2,0),(2,-1),colors.HexColor('#E8F1F8')),('BACKGROUND',(4,0),(4,-1),colors.HexColor('#E8F1F8')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('PADDING',(0,0),(-1,-1),3)])); story += [t,Spacer(1,6)]
    subjects=[]
    try: subjects=list(json.loads(student.get('Subject_Data','{}')).keys())
    except: pass
    if not subjects: subjects=[s for s in XI_SUBJECTS+IX_SUBJECTS+LOWER_SUBJECTS if s in student]
    try: bd=json.loads(student.get('Breakdown','{}'))
    except: bd={}
    rows=[['SUBJECT','TERM-I','TERM-II','FINAL','GRADE']]
    for s in subjects:
        if s not in student: continue
        b=bd.get(s,{}) if isinstance(bd,dict) else {}; final=num(student.get(s)); rows.append([s,str(b.get('term1','') if b.get('term1') is not None else ''),str(b.get('term2','') if b.get('term2') is not None else ''),f'{final:.2f}' if not pd.isna(final) else '',grade(final)])
    rows.append(['OVERALL','','',f"{num(student.get('Total_Marks')):.2f}",grade(num(student.get('Percentage')))])
    stt=Table(rows,colWidths=[160,80,80,90,70],repeatRows=1); stt.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.5,colors.HexColor('#555555')),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#003366')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('ALIGN',(1,0),(-1,-1),'CENTER'),('FONTSIZE',(0,0),(-1,-1),7),('PADDING',(0,0),(-1,-1),3)])); story += [stt,Spacer(1,6)]
    summary=[['Total Marks',f"{num(student.get('Total_Marks')):.2f} / {num(student.get('Max_Marks')):.0f}"],['Percentage',f"{num(student.get('Percentage')):.2f}%"],['Class Rank',norm(student.get('Class_Rank'))],['Attendance',norm(student.get('Attendance'))],['Discipline',norm(student.get('Discipline'))],['Skill Course',norm(student.get('Skill_Course'))],['Achievement',norm(student.get('Outstanding_Achievement'))],['Remarks',norm(student.get('Remarks'))]]
    st2=Table(summary,colWidths=[100,380]); st2.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.4,colors.grey),('BACKGROUND',(0,0),(0,-1),colors.HexColor('#F1F6F1')),('FONTSIZE',(0,0),(-1,-1),7),('PADDING',(0,0),(-1,-1),3)])); story += [st2,Spacer(1,18),Table([['Class Teacher Signature','Principal Signature']],colWidths=[240,240],rowHeights=[30]),Paragraph('Generated from the latest imported/edited student database. Previous values are preserved in update history.',center)]
    doc.build(story); return buf.getvalue()

st.set_page_config(page_title=APP_TITLE,page_icon='🎓',layout='wide')
st.title('🎓 PM SHRI JNV CHHOTAUDEPUR — RESULT PORTAL')
st.caption('Excel-compatible import • old data preserved • new updates merged • individual & bulk report cards')

if 'df' not in st.session_state: st.session_state.df=db_load()

with st.sidebar:
    st.header('📁 Excel Import')
    uploads=st.file_uploader('Upload one or more result Excel files',type=['xlsx'],accept_multiple_files=True)
    if st.button('🔄 Import + Merge Excel Data',use_container_width=True):
        if not uploads: st.warning('Please upload Excel files first.')
        else:
            backup_db(); all_df=[]; errors=[]
            for up in uploads:
                try:
                    temp=os.path.join('/tmp',up.name); open(temp,'wb').write(up.getbuffer()); d=import_template(temp)
                    if d.empty: errors.append(up.name+': no student rows found')
                    else: all_df.append(d)
                except Exception as e: errors.append(up.name+': '+str(e))
            if all_df:
                merged=pd.concat(all_df,ignore_index=True,sort=False); db_merge(merged); st.session_state.df=db_load(); st.success(f'{len(merged)} Excel rows imported/merged.')
            for e in errors: st.error(e)
    st.divider()
    if st.button('💾 Create DB Backup',use_container_width=True): backup_db(); st.success('Backup created.')

DF=st.session_state.df
if DF.empty:
    st.info('Upload the supplied VI–VIII, IX or XI Excel file(s) from the sidebar. The portal will keep existing database values and update only non-empty new values.')
    st.stop()

c1,c2,c3,c4=st.columns(4)
c1.metric('Students',DF[['Class','Roll_No']].drop_duplicates().shape[0])
c2.metric('Classes',DF['Class'].nunique())
c3.metric('Highest %',f"{pd.to_numeric(DF['Percentage'],errors='coerce').max():.2f}%")
c4.metric('Average %',f"{pd.to_numeric(DF['Percentage'],errors='coerce').mean():.2f}%")

tab1,tab2,tab3,tab4=st.tabs(['🔎 Student Result','📥 All Report Cards','✏️ Edit / Update','🕘 Update History'])
with tab1:
    classes=sorted(DF['Class'].astype(str).unique()); cls=st.selectbox('Class',classes)
    sub=DF[DF['Class'].astype(str)==cls].copy(); rolls=sorted(sub['Roll_No'].astype(str).unique(),key=lambda x:(len(x),x)); roll=st.selectbox('Roll No.',rolls)
    s=sub[sub['Roll_No'].astype(str)==roll].iloc[0]
    st.subheader(f"{s.get('Student_Name','')} — Class {cls}, Roll {roll}")
    st.dataframe(pd.DataFrame({'Field':['Father','Mother','DOB','House','Area','Attendance','Discipline','Skill Course','Total','Percentage','Rank'], 'Value':[s.get('Father_Name',''),s.get('Mother_Name',''),s.get('DOB',''),s.get('House',''),s.get('Area',''),s.get('Attendance',''),s.get('Discipline',''),s.get('Skill_Course',''),s.get('Total_Marks',''),f"{num(s.get('Percentage')):.2f}%",s.get('Class_Rank','')] }),hide_index=True,use_container_width=True)
    try: subjects=list(json.loads(s.get('Subject_Data','{}')).keys())
    except: subjects=[]
    marks=[{'Subject':x,'Final':num(s.get(x)),'Grade':grade(num(s.get(x)))} for x in subjects if x in s]
    if marks: st.dataframe(pd.DataFrame(marks),hide_index=True,use_container_width=True)
    st.download_button('📄 Download Report Card PDF',make_report(s),f"Report_{cls}_{roll}.pdf",'application/pdf')

with tab2:
    st.write('Generate one ZIP containing **every student report card** in the selected class, or all classes.')
    mode=st.radio('Scope',['Selected Class','All Classes'],horizontal=True)
    selected=st.selectbox('Select Class',sorted(DF['Class'].astype(str).unique())) if mode=='Selected Class' else None
    if st.button('🚀 Generate All Report Cards',use_container_width=True):
        target=DF if selected is None else DF[DF['Class'].astype(str)==selected]
        bio=io.BytesIO()
        with zipfile.ZipFile(bio,'w',zipfile.ZIP_DEFLATED) as z:
            seen=set()
            for _,r in target.iterrows():
                k=key_for(r)
                if k in seen: continue
                seen.add(k); z.writestr(f"Class_{r['Class']}_Roll_{r['Roll_No']}_{re.sub(r'[^A-Za-z0-9_-]','_',norm(r.get('Student_Name')))}.pdf",make_report(r))
        bio.seek(0); st.download_button('📦 Download ZIP — All Report Cards',bio,f"All_Report_Cards_{selected or 'ALL'}.zip",'application/zip',use_container_width=True)

with tab3:
    st.write('Edit a student. Changes are saved separately and the previous value is preserved in the history table.')
    cls=st.selectbox('Edit Class',sorted(DF['Class'].astype(str).unique()),key='edit_cls'); sub=DF[DF['Class'].astype(str)==cls]; roll=st.selectbox('Edit Roll No.',sorted(sub['Roll_No'].astype(str).unique()),key='edit_roll'); idx=DF[(DF['Class'].astype(str)==cls)&(DF['Roll_No'].astype(str)==roll)].index[0]; s=DF.loc[idx]
    with st.form('edit_form'):
        name=st.text_input('Student Name',norm(s.get('Student_Name'))); father=st.text_input('Father Name',norm(s.get('Father_Name'))); mother=st.text_input('Mother Name',norm(s.get('Mother_Name'))); house=st.text_input('House',norm(s.get('House'))); area=st.text_input('Area',norm(s.get('Area'))); mobile=st.text_input('Mobile',norm(s.get('Mobile_No'))); attendance=st.text_input('Attendance',norm(s.get('Attendance'))); discipline=st.selectbox('Discipline',['A','B','C',''],index=['A','B','C',''].index(norm(s.get('Discipline'))) if norm(s.get('Discipline')) in ['A','B','C',''] else 0); remarks=st.text_input('Remarks',norm(s.get('Remarks'))); achievement=st.text_input('Outstanding Achievement',norm(s.get('Outstanding_Achievement')))
        save=st.form_submit_button('💾 Save Modification',use_container_width=True)
    if save:
        old=dict(s); updates={'Class':cls,'Roll_No':roll,'Student_Name':name,'Father_Name':father,'Mother_Name':mother,'House':house,'Area':area,'Mobile_No':mobile,'Attendance':attendance,'Discipline':discipline,'Remarks':remarks,'Outstanding_Achievement':achievement}; new=old.copy(); new.update(updates); con=sqlite3.connect(DB_FILE); now=datetime.now().isoformat(timespec='seconds'); con.execute('INSERT INTO history(key,action,old_data,new_data,changed_at) VALUES(?,?,?,?,?)',(key_for(old),'MANUAL_EDIT',json.dumps(old,default=str),json.dumps(new,default=str),now)); con.execute('UPDATE students SET data=?,updated_at=? WHERE key=?',(json.dumps(new,default=str,ensure_ascii=False),now,key_for(old))); con.commit(); con.close(); st.session_state.df=db_load(); st.success('Updated successfully. Old data is preserved in Update History.'); st.rerun()

with tab4:
    con=sqlite3.connect(DB_FILE); hist=pd.read_sql_query('SELECT id,key,action,changed_at FROM history ORDER BY id DESC',con); con.close(); st.dataframe(hist,hide_index=True,use_container_width=True)
    st.caption('Every Excel import and manual edit creates a history entry before replacing the current value.')

st.divider(); st.caption('The supplied templates use different layouts for VI–VIII/IX and XI. This version reads their DATA/DataSheet structures, recalculates totals/ranks in Python, merges new non-empty values with existing database records, and uses the merged data for report cards.')
